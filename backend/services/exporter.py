import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def generate_excel_export(session_title: str, leads: list) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Лид-трекер"

    # Title Banner
    ws.merge_cells("A1:J1")
    title_cell = ws["A1"]
    title_cell.value = f"Лид-трекер: {session_title}"
    title_cell.font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    headers = [
        "№",
        "Профиль Instagram",
        "Имя / Название",
        "Ссылка",
        "Подписчики",
        "Наличие сайта / Статус",
        "Внешняя ссылка",
        "Писал? (Контакт)",
        "Статус ответа",
        "Био / Описание"
    ]

    ws.append([]) # Row 2 empty spacer
    ws.append(headers) # Row 3
    ws.row_dimensions[3].height = 24

    # Header styling
    header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )

    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=3, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    # Lead Rows
    for idx, lead in enumerate(leads, start=1):
        row_num = idx + 3
        contacted_text = "ДА" if lead.contacted else "НЕТ"
        
        row_data = [
            idx,
            f"@{lead.username}",
            lead.full_name or "-",
            lead.profile_url,
            lead.followers_count,
            lead.link_label,
            lead.external_url or "—",
            contacted_text,
            lead.reply_status or "Не отправлено",
            lead.biography or ""
        ]
        ws.append(row_data)
        ws.row_dimensions[row_num].height = 20

        # Row styling
        # Highlight "Нет сайта" leads with soft red/rose tint
        is_no_site = lead.link_type == "no_site"
        row_fill = PatternFill(start_color="FFF1F2" if is_no_site else "FFFFFF", fill_type="solid")

        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_num, column=col_idx)
            cell.font = Font(name="Calibri", size=10)
            cell.border = thin_border
            cell.fill = row_fill
            if col_idx in [1, 5, 8]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

    # Auto-fit column widths
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row < 3: # Skip title banner
                continue
            val_str = str(cell.value or '')
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 40)

    # Specific fixed widths for key columns
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 24
    ws.column_dimensions['C'].width = 26
    ws.column_dimensions['D'].width = 32
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 24
    ws.column_dimensions['G'].width = 28
    ws.column_dimensions['H'].width = 14
    ws.column_dimensions['I'].width = 20
    ws.column_dimensions['J'].width = 40

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream